# -*- coding: utf-8 -*-
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

with open(r"C:\finances\data\sba_loans\_singlebid_detail_export.json", encoding="utf-8") as f:
    records = json.load(f)

COLUMNS = [
    ("award_id", "Award ID", 16),
    ("parent_award_id", "Parent Award ID", 16),
    ("vendor", "Vendor", 32),
    ("vendor_uei", "Vendor UEI", 14),
    ("amount", "Amount ($)", 13),
    ("near_threshold", "Near Threshold ($300K-$349,999)", 14),
    ("branch_short", "Branch", 10),
    ("awarding_sub_agency", "Awarding Sub-Agency", 24),
    ("awarding_office", "Awarding Office", 26),
    ("funding_agency", "Funding Agency", 20),
    ("funding_sub_agency", "Funding Sub-Agency", 24),
    ("extent_competed", "Extent Competed", 30),
    ("solicitation_procedures", "Solicitation Procedures", 22),
    ("solicitation_id", "Solicitation ID", 18),
    ("type_of_contract_pricing", "Contract Pricing Type", 18),
    ("number_of_offers", "Offers Received", 10),
    ("award_date", "Award Date", 12),
    ("pop_start_date", "Period of Perf. Start", 14),
    ("pop_end_date", "Period of Perf. End", 14),
    ("description", "Description / Memo", 50),
    ("psc_description", "Product/Service Code", 40),
    ("naics_description", "NAICS Description", 40),
    ("pop_city", "Place of Performance - City", 20),
    ("pop_state", "Place of Performance - State", 16),
    ("pop_country", "Place of Performance - Country", 16),
    ("vendor_city", "Vendor City", 18),
    ("vendor_state", "Vendor State", 14),
    ("vendor_country", "Vendor Country", 14),
    ("place_of_manufacture", "Place of Manufacture", 22),
]

wb = Workbook()
ws = wb.active
ws.title = "Single-Bid Under $350K"

FONT_NAME = "Arial"
header_font = Font(name=FONT_NAME, size=10, bold=True, color="FFFFFF")
header_fill = PatternFill("solid", fgColor="142433")
body_font = Font(name=FONT_NAME, size=10)
near_fill = PatternFill("solid", fgColor="F1E6FA")
thin = Side(style="thin", color="D9D9D9")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# Header row
for col_idx, (key, label, width) in enumerate(COLUMNS, start=1):
    cell = ws.cell(row=1, column=col_idx, value=label)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(vertical="center", wrap_text=True)
    cell.border = border
    ws.column_dimensions[get_column_letter(col_idx)].width = width
ws.row_dimensions[1].height = 30

# Data rows
for row_idx, r in enumerate(records, start=2):
    for col_idx, (key, label, width) in enumerate(COLUMNS, start=1):
        val = r.get(key)
        if key == "near_threshold":
            val = "Yes" if val else "No"
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.font = body_font
        cell.border = border
        if key == "amount":
            cell.number_format = "$#,##0"
        if key in ("description", "psc_description", "naics_description"):
            cell.alignment = Alignment(vertical="top", wrap_text=False)
        if r.get("near_threshold"):
            cell.fill = near_fill

ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{len(records)+1}"

# Summary sheet
ws2 = wb.create_sheet("Summary")
ws2.column_dimensions["A"].width = 46
ws2.column_dimensions["B"].width = 20
summary_rows = [
    ("Scope", "DoD FY2025 definitive contracts (award type D), single-bid, genuinely-competed, under $350,000 SAT"),
    ("Total contracts", len(records)),
    ("Total value", sum(r["amount"] for r in records)),
    ("Near-threshold ($300K-$349,999) contracts", sum(1 for r in records if r["near_threshold"])),
    ("Near-threshold value", sum(r["amount"] for r in records if r["near_threshold"])),
    ("", ""),
    ("Source", "USAspending.gov bulk award-download API, queried live"),
    ("Extent Competed definition", "Full and Open Competition, Full and Open Competition After Exclusion of Sources, or Competed Under SAP (sole-source categories excluded)"),
    ("Branch", "Awarding sub-agency as reported on the award; not necessarily the end-use branch"),
    ("Description / Memo", "prime_award_base_transaction_description field from USAspending's bulk export"),
]
for i, (label, val) in enumerate(summary_rows, start=1):
    lc = ws2.cell(row=i, column=1, value=label)
    lc.font = Font(name=FONT_NAME, size=10, bold=True)
    vc = ws2.cell(row=i, column=2, value=val)
    vc.font = Font(name=FONT_NAME, size=10)
    vc.alignment = Alignment(wrap_text=True, vertical="top")
    if "value" in label.lower() and isinstance(val, (int, float)):
        vc.number_format = "$#,##0"

wb.save(r"C:\finances\data\sba_loans\DoD_SingleBid_Contracts_Under_350K.xlsx")
print("Saved.", len(records), "rows")
